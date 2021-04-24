import pandas as pd
from openpyxl.reader.excel import load_workbook
from tkinter.filedialog import askopenfilename


def pn_consolidado_header(name_file):
    df = pd.DataFrame(columns=['linha_id', 'Tipo PN', 'Data de Inclusão',
                               'Regional', 'Hostname', 'Site (BDRaf)',
                               'SWAP DE', 'Projeto', 'Sub Projeto', 'Anel IP RAN', 'IP NODEB', 'Integrado',
                               'Fabricante', 'Integrador', 'Aplicação', 'Modelo Bastidor', 'Modelo Roteador Simplificado (Inicial)',
                               'Municipio', 'Remanejamento', 'Comentário', 'Projeto CAPEX', 'Sub-Projeto CAPEX', 'SI CAPEX', 'Ano Implantação', 'Executado'])

    writer = pd.ExcelWriter(name_file, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='PNs Consolidados', index=False)
    writer.save()
    return


def le_pn_ipran():
    # Dados fixos
    tipo_pn = 'IPRAN'
    # Dados base orçamentária
    filename = askopenfilename()
    df1 = pd.read_excel(filename, sheet_name='Relatorio 1')
    df1['filtro'] = df1['Fabricante'].apply(lambda row: row.upper())
    df1 = df1[(df1['filtro'] == "CISCO")]
    line_id = df1['linha_id'].to_list()
    data_inc = df1['Data de Inclusão'].to_list()
    regional = df1['Regional'].to_list()
    hostname = df1['Hostname'].to_list()
    site = df1['Site (BDRaf)'].to_list()
    swap_de = df1['SWAP DE'].to_list()
    projeto = df1['Projeto'].to_list()
    sub_projeto = df1['Sub Projeto'].to_list()
    anel_ipran = df1['Anel IP RAN'].to_list()
    anel_ipnodeb = df1['IP NODEB'].to_list()
    integrado = df1['Integrado'].to_list()
    fabricante = df1['Fabricante'].to_list()
    integrador = df1['Integrador'].to_list()
    aplicacao = df1['Aplicação'].to_list()
    model_bastidor = df1['Modelo Bastidor'].to_list()
    model_simp = df1['Modelo Roteador Simplificado (Inicial)'].to_list()
    municipio = df1['Municipio'].to_list()
    remanejamento = df1['Remanejamento'].to_list()
    comentario = df1['Comentário'].to_list()
    projeto_capex = df1['Projeto CAPEX'].to_list()
    sub_projeto_capex = df1['Sub-Projeto CAPEX'].to_list()
    si_capex = df1['SI CAPEX'].to_list()
    ano_imp = df1['Ano Implantação'].to_list()

    return tipo_pn, line_id, data_inc, regional, hostname, site, swap_de, projeto,\
        sub_projeto, anel_ipran, anel_ipnodeb, integrado, fabricante, integrador, aplicacao,\
        model_bastidor, model_simp, municipio, remanejamento, comentario,\
        projeto_capex, sub_projeto_capex, si_capex, ano_imp


def pn_consolidado_escreve_ipran(name_file, tipo_pn, line_id, data_inc, regional, hostname, site, swap_de, projeto,
                                 sub_projeto, anel_ipran, anel_ipnodeb, integrado, fabricante, integrador, aplicacao,
                                 model_bastidor, model_simp, municipio, remanejamento, comentario,
                                 projeto_capex, sub_projeto_capex, si_capex, ano_imp):

    df = pd.read_excel(name_file, sheet_name='PNs Consolidados')
    df['linha_id'] = line_id
    df['Tipo PN'] = tipo_pn
    df['Data de Inclusão'] = data_inc
    df['Regional'] = regional
    df['Hostname'] = hostname
    df['Site (BDRaf)'] = site
    df['SWAP DE'] = swap_de
    df['Projeto'] = projeto
    df['Sub Projeto'] = sub_projeto
    df['Anel IP RAN'] = anel_ipran
    df['IP NODEB'] = anel_ipnodeb
    df['Integrado'] = integrado
    df['Fabricante'] = fabricante
    df['Integrador'] = integrador
    df['Aplicação'] = aplicacao
    df['Modelo Bastidor'] = model_bastidor
    df['Modelo Roteador Simplificado (Inicial)'] = model_simp
    df['Municipio'] = municipio
    df['Remanejamento'] = remanejamento
    df['Comentário'] = comentario
    df['Projeto CAPEX'] = projeto_capex
    df['Sub-Projeto CAPEX'] = sub_projeto_capex
    df['SI CAPEX'] = si_capex
    df['Ano Implantação'] = ano_imp
    df['Executado'] = '-'

    df.to_excel(name_file, sheet_name='PNs Consolidados', index=False)
    return


def le_pn_upgrade():
    # Dados fixos
    tipo_pn = 'Upgrade'
    # Dados base orçamentária
    filename = askopenfilename()
    df1 = pd.read_excel(filename, sheet_name='Relatorio 1')
    df1['filtro'] = df1['Integrador'].apply(lambda row: str(row).upper())
    df1 = df1[(df1['filtro'] == "PROMON")]
    line_id = df1['linha_id'].to_list()
    data_inc = df1['Data de Inclusão'].to_list()
    regional = df1['Regional'].to_list()
    hostname = df1['Hostname'].to_list()
    site = df1['Site (BDRaf)'].to_list()
    projeto = df1['Projeto'].to_list()
    sub_projeto = df1['Sub Projeto'].to_list()
    anel_ipran = df1['Anel IP RAN'].to_list()
    anel_ipnodeb = df1['IP NODEB'].to_list()
    integrado = df1['Integrado'].to_list()
    fabricante = df1['Fabricante'].to_list()
    aplicacao = df1['Aplicação'].to_list()
    model_bastidor = df1['Modelo Bastidor'].to_list()
    comentario = df1['Comentário'].to_list()
    projeto_capex = df1['Projeto CAPEX'].to_list()
    sub_projeto_capex = df1['Sub-Projeto CAPEX'].to_list()
    si_capex = df1['SI CAPEX'].to_list()
    ano_imp = df1['Ano Implantação'].to_list()
    executado = df1['Executado'].to_list()
    integrador = df1['Integrador'].to_list()

    return tipo_pn, line_id, data_inc, regional, hostname, site, projeto,\
        sub_projeto, anel_ipran, anel_ipnodeb, integrado, fabricante, integrador, aplicacao,\
        model_bastidor, comentario, projeto_capex, sub_projeto_capex, si_capex, ano_imp, executado


def pn_consolidado_escreve_upgrade(name_file, tipo_pn, line_id, data_inc, regional, hostname, site, projeto,
                                   sub_projeto, anel_ipran, anel_ipnodeb, integrado, fabricante, integrador, aplicacao,
                                   model_bastidor, comentario, projeto_capex, sub_projeto_capex, si_capex, ano_imp, executado):
    df = pd.DataFrame()
    df['linha_id'] = line_id
    df['Tipo PN'] = tipo_pn
    df['Data de Inclusão'] = data_inc
    df['Regional'] = regional
    df['Hostname'] = hostname
    df['Site (BDRaf)'] = site
    df['SWAP DE'] = '-'
    df['Projeto'] = projeto
    df['Sub Projeto'] = sub_projeto
    df['Anel IP RAN'] = anel_ipran
    df['IP NODEB'] = anel_ipnodeb
    df['Integrado'] = integrado
    df['Fabricante'] = fabricante
    df['Integrador'] = integrador
    df['Aplicação'] = aplicacao
    df['Modelo Bastidor'] = model_bastidor
    df['Modelo Roteador Simplificado (Inicial)'] = '-'
    df['Municipio'] = '-'
    df['Remanejamento'] = '-'
    df['Comentário'] = comentario
    df['Projeto CAPEX'] = projeto_capex
    df['Sub-Projeto CAPEX'] = sub_projeto_capex
    df['SI CAPEX'] = si_capex
    df['Ano Implantação'] = ano_imp
    df['Executado'] = executado

    book = load_workbook(name_file)
    writer = pd.ExcelWriter(name_file, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    startrow1 = writer.book['PNs Consolidados'].max_row

    df.to_excel(writer, sheet_name='PNs Consolidados',
                index=False, startrow=startrow1, header=False)
    writer.save()
    return
